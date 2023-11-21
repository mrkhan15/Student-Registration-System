from customtkinter import *
from customtkinter.windows import CTk
from customtkinter import CTkLabel
import tkinter as tk
import pathlib
from datetime import date
from tkinter import filedialog, PhotoImage, messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook, load_workbook
from customtkinter import CTkFont



app = CTk()
app.title("Student Registration System")
# set_appearance_mode("System")
app.geometry("500x400")

# Activate this if using in Monitor screen "Set your own scaling factor for widget dimensions and text size"
# set_widget_scaling(1.5)


#Save Data to Exel file
file=pathlib.Path('Test_Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='Registration No.'
    sheet['B1']='Full Name'
    sheet['C1']='Date of Birth'
    sheet['D1']='Degree'
    sheet['E1']='Major'
    sheet['F1']='Email Address'
    sheet['G1']='Contact Number'
    sheet['H1']='Address'
    sheet['I1']="Father's Name"
    sheet['J1']='Previous School/ Instituion'
    sheet['K1']='Gender'

    file.save('Test_Data.xlsx')


# Adjust family and size as needed
font = CTkFont(family='Roboto', size=15, weight='bold')  

#Top_Frames 
label = CTkLabel(app, text="STUDENT REGISTRATION", font=font, width=2, height=50, fg_color='#c0c9fe')
label.pack(fill=X)

# main_view
main_view = CTkFrame(master=app,  corner_radius=0,)
main_view.pack(side="top")

#Clear Button function 
def clear():
    Name.set('')
    DOB.set('')
    Degree.set('')
    Major.set("")
    Email.set('')
    Contact.set('')
    Address.set('')
    FatherName.set('')
    PrvSchool.set('')
    # radiobutton_event.set(None)

    # Savebutton.configure(state='normal')

    # Reload default image
    default_image = PhotoImage(file="Images/upload holder.png")
    profile_label.configure(image=default_image)
    profile_label.image = default_image

# Registration -- It will check the data of last row and add 1 to the reg no.
def registration_no(registration_var):
    try:
        file = openpyxl.load_workbook('Test_Data.xlsx')
        sheet = file.active
        row = sheet.max_row
        max_row_value = sheet.cell(row=row, column=1).value
        reg_value = int(max_row_value) + 1
        registration_var.set(reg_value)
        file.close()  # Close the workbook
    except Exception as e:
        print(f"An error occurred: {e}")
        registration_var.set('1')


CTkLabel(app, text="Registration No:", font=font).place(x=30, y=150)

Registration = StringVar()
reg_entry = CTkEntry(app, textvariable=Registration, width=80, font=font)
reg_entry.place(x=160, y=150)

registration_no(Registration)

# Create a button to trigger the search action

def perform_search():
    try:
        text = search_entry.get()
        clear()
        print(f"Performing search for: {text}")
        Savebutton.configure(state='disable')

        file = load_workbook("Test_Data.xlsx")
        sheet = file.active

        # Find the column index for "Registration No." header
        registration_number_col = None
        for col_index, col in enumerate(sheet.iter_cols(min_row=1, max_row=1), 1):
            if col[0].value == "Registration No.":
                registration_number_col = col_index
                break

        if registration_number_col is not None:
            print(f"Found 'Registration number' column at index: {registration_number_col}")

            # Iterate over rows starting from the second row
            for row in sheet.iter_rows(min_row=2):
                reg_number_cell = row[registration_number_col - 1]  # Adjusting for 0-based index
                if str(reg_number_cell.value) == str(text):
                    # Assuming you have labels for displaying data
                    Name.set(row[1].value)   # Full Name
                    DOB.set(row[2].value)    # Date of Birth
                    Degree.set(row[3].value)  # Degree
                    Major.set(row[4].value)   # Major
                    Email.set(row[5].value)   # Email
                    Contact.set(row[6].value)  # Contact
                    Address.set(row[7].value)  # Address
                    FatherName.set(row[8].value)  # Father's Name
                    PrvSchool.set(row[9].value)   # Previous School
                    # ... (Update other fields similarly)

                    print(f"Data found for registration number {text}")
                    return  # Exit the loop once data is found

            # If the loop completes without finding data
            print(f"No data found for the given registration number {text}")
            messagebox.showerror("Error", f"No data found for the given registration number {text}")
        else:
            print("Column 'Registration number' not found in the Excel sheet")
            messagebox.showerror("Error", "Column 'Registration number' not found in the Excel sheet")

    except FileNotFoundError as e:
        print(f"File not found: {e}")
        messagebox.showerror("Error", "File not found")
    except Exception as e:
        print(f"An error occurred: {e}")
        messagebox.showerror("Error", "An error occurred while searching")


# Search bar
search_container = CTkFrame(master=main_view, height=50, fg_color="#F0F0F0")
search_container.pack(fill="x")

search_entry = CTkEntry(master=search_container, width=305, placeholder_text="Search", border_color="#c0c9fe", border_width=2)
search_entry.pack(side="left")

search_button_image = PhotoImage(file="search_icon.png")
search_button = CTkButton(search_container, image=search_button_image, text=" ", fg_color='#c0c9fe', width=20, height=20, command=perform_search)
search_button.pack(side="left")

#Date --------->
 
Date=StringVar()
CTkLabel(app, text="Date:", font=font).place(x=500,y=150)
today=date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = CTkEntry(app, textvariable=Date, width=80, font=font)
date_entry.place(x=550, y=150)
Date.set(d1)

#Student Details ------>

label2 = CTkLabel(app, font=font,text="", width=900, height=250,  fg_color='#c0c9fe')
label2.place(x=30, y=200)

CTkLabel(label2, text="Full Name:", font=font, ).place(x=30,y=50)

Name=StringVar()
name_entry = CTkEntry(label2, textvariable=Name, width=150, font=font)
name_entry.place(x=160, y=50)

CTkLabel(label2, text="Date of Birth:", font=font, ).place(x=30,y=100)
DOB=StringVar()
dob_entry = CTkEntry(label2, textvariable=DOB, width=150, font=font)
dob_entry.place(x=160, y=100)

CTkLabel(label2, text="Gender:", font=font, ).place(x=30,y=150) #Gender:

#Gender Selection
def radiobutton_event():
    global gender_value
    value = radio.get()
    if value == 1:
        gender="Male"
        print (gender)
    else:
        gender="Female"
        print(gender)

radio=IntVar()
R1 = CTkRadioButton(label2, text="Male", variable=radio, value=1, command=radiobutton_event)
R1.place(x=150,y=150)

R2 = CTkRadioButton(label2, text="Female", variable=radio,value=2,command=radiobutton_event)
R2.place(x=250,y=150)
##########

CTkLabel(label2, text="Degree:", font=font, ).place(x=500,y=50)

Degree=StringVar()
Degree_entry = CTkEntry(label2, textvariable=Degree, width=150, font=font)
Degree_entry.place(x=630, y=50)


CTkLabel(label2, text="Major:", font=font, ).place(x=500,y=100)

Major=StringVar()
Major_entry = CTkEntry(label2, textvariable=Major, width=150, font=font)
Major_entry.place(x=630, y=100)


CTkLabel(label2, text="Email Address:", font=font, ).place(x=500,y=150)

Email=StringVar()
Email_entry = CTkEntry(label2, textvariable=Email, width=150, font=font)
Email_entry.place(x=630, y=150)

#Other Details

label3 = CTkLabel(app, font=font,text="", width=900, height=250,  fg_color='#c0c9fe')
label3.place(x=30, y=470)


CTkLabel(label3, text="Contact Number:", font=font, ).place(x=30,y=50)

Contact=StringVar()
Contact_entry = CTkEntry(label3, textvariable=Contact, width=150, font=font)
Contact_entry.place(x=160, y=50)

CTkLabel(label3, text="Address:", font=font, ).place(x=30,y=100)

Address=StringVar()
Address_entry = CTkEntry(label3, textvariable=Address, width=150, font=font)
Address_entry.place(x=160, y=100)

CTkLabel(label3, text="Father's Name:", font=font, ).place(x=500,y=50)

FatherName=StringVar()
FatherName_entry = CTkEntry(label3, textvariable=FatherName, width=150, font=font)
FatherName_entry.place(x=630, y=50)

CTkLabel(label3, text="Previous School:", font=font, ).place(x=500,y=100)

PrvSchool=StringVar()
PrvSchool_entry = CTkEntry(label3, textvariable=PrvSchool, width=150, font=font)
PrvSchool_entry.place(x=630, y=100)

#Image Section

f=CTkFrame(app, width=200, height=200)
f.place(x=1000, y=150)

profile_image = PhotoImage(file="Images/upload holder.png")
profile_label = CTkLabel(f, text='', image=profile_image, width=200, height=200)
profile_label.place(x=0, y=0)

#Show Image

def show_image():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select the Image File",
        filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All Files", "*.*"))
    )
    
    if filename:        
        img = Image.open(filename)
        resized_image =img.resize((300, 300))
        photo2 = ImageTk.PhotoImage(resized_image)
        profile_label.configure(image=photo2)
        # label.configure(image=photo2)
        label.image = photo2

# profile_label = CTkLabel(f, text='')
profile_label.place(x=0, y=0)

#Buttons------------------->

upload_button = CTkButton(app, text="Upload", corner_radius=32, fg_color='#2A409A', hover_color='#c0c9fe', command=show_image)
upload_button.place(x=1000, y=370)

def Save():
    R1=Registration.get()
    N1=Name.get()
    B1=DOB.get()
    D1=Degree.get()
    M1=Major.get()
    E1=Email.get()
    C1=Contact.get()
    A1=Address.get()
    F1=FatherName.get()
    P1=PrvSchool.get()

    if R1=='' or N1=='' or D1=='' or B1=='' or M1=='' or E1=='' or C1=='' or A1=='' or F1=='' or P1=='':
        messagebox.showerror("ERROR!", "Few Data is missing.")
    else:
        file=openpyxl.load_workbook('Test_Data.xlsx')
        sheet=file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=B1)
        sheet.cell(column=4, row=sheet.max_row, value=D1)
        sheet.cell(column=5, row=sheet.max_row, value=M1)
        sheet.cell(column=6, row=sheet.max_row, value=E1)
        sheet.cell(column=7, row=sheet.max_row, value=C1)
        sheet.cell(column=8, row=sheet.max_row, value=A1)
        sheet.cell(column=9, row=sheet.max_row, value=F1)
        sheet.cell(column=10, row=sheet.max_row, value=P1)

        file.save('Test_Data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".png")
        except:
            messagebox.showinfo("INFO!", "Profile picture is not available")

        messagebox.showinfo("SUCESSFUL!", "Data has been stored Sucessfully")

        registration_no(Registration)


        clear() #Clear the entire form after Saving the Data.


########### Need to fix The gender check before wined up.!!!!!!!!!!!!!!!
    # if radio.get() == 0:
    #     messagebox.showerror("ERROR!", "Please Select Gender")
    # print(R1)

### Sea  


Savebutton = CTkButton(app, text="Save", corner_radius=32, fg_color='green', hover_color='#c0c9fe', command=Save)
Savebutton.place(x=1000, y=450)

#Clear Button     
Resetbutton = CTkButton(app, text="Reset", corner_radius=32, fg_color='grey',hover_color='#c0c9fe', command=clear)
Resetbutton.place(x=1000, y=530)

#EXIT Switch
def exit():
    app.destroy()

Exitbutton = CTkButton(app, text="Exit", command=exit, corner_radius=32, fg_color='red', hover_color='#c0c9fe' )
Exitbutton.place(x=1000, y=610)



app.mainloop()
