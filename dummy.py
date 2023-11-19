from customtkinter import *
from customtkinter.windows import CTk
from customtkinter import CTkLabel
import tkinter as tk
import pathlib
from datetime import date
from tkinter import filedialog, PhotoImage, messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
from customtkinter import CTkFont



app = CTk()
app.title("Student Registration System")
# set_appearance_mode("System")
app.geometry("500x400")



#Save Data to Exel file
file=pathlib.Path('Dummy_data4.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='Registration No.'
    sheet['B1']='Full Name'
    sheet['C1']='Date of Birth'
    sheet['D1']='Gender'
    sheet['E1']='Degree'
    sheet['F1']='Major'
    sheet['G1']='Email Address'
    sheet['H1']='Contact Number'
    sheet['I1']='Address'
    sheet['J1']="Father's Name"
    sheet['K1']='Parent / Guardian Contact No.'
    sheet['L1']='Previous School/ Instituion'

    file.save('Dummy_data4.xlsx')


# Adjust family and size as needed
font = CTkFont(family='Roboto', size=15, weight='bold')  

#Top_Frames 
label = CTkLabel(app, text="STUDENT REGISTRATION", font=font, width=2, height=50, fg_color='#c0c9fe')
label.pack(fill=X)

# main_view
main_view = CTkFrame(master=app,  corner_radius=0,)
main_view.pack(side="top")


# Create a button to trigger the search action
def perform_search():
    query = search_entry.get()
    # Add your search logic here
    print(f"Performing search for: {query}")

# Search bar
search_container = CTkFrame(master=main_view, height=50, fg_color="#F0F0F0")
search_container.pack(fill="x")

search_entry = CTkEntry(master=search_container, width=305, placeholder_text="Search", border_color="#c0c9fe", border_width=2)
search_entry.pack(side="left")

search_button_image = PhotoImage(file="search_icon.png")
search_button = CTkButton(search_container, image=search_button_image, text=" ", fg_color='#c0c9fe', width=20, height=20, command=perform_search)
search_button.pack(side="left")

#Registration -- It will check the data of last row and add 1 to the reg no.

def registration_no():
    file=openpyxl.load_workbook('Dummy_data4.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_Value=sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_Value+1)
    except:
        Registration.set('1')
    
CTkLabel(app, text="Registration No:", font=font, ).place(x=30,y=150)

Registration=StringVar()
reg_entry= CTkEntry(app, textvariable=Registration, width=80, font=font)
reg_entry.place(x=160, y=150)

registration_no()


#Date 
Date=StringVar()
CTkLabel(app, text="Date:", font=font).place(x=500,y=150)
today=date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = CTkEntry(app, textvariable=Date, width=80, font=font)
date_entry.place(x=550, y=150)
Date.set(d1)

#Student Details

label2 = CTkLabel(app, font=font,text="", width=900, height=250,  fg_color='#c0c9fe')
label2.place(x=30, y=200)

CTkLabel(label2, text="Full Name:", font=font, ).place(x=30,y=50)

Name=StringVar()
name_entry = CTkEntry(label2, textvariable=Name, width=150, font=font)
name_entry.place(x=160, y=50)

CTkLabel(label2, text="Date of Birth:", font=font, ).place(x=30,y=100)
DOB=StringVar()
dob_entry = CTkEntry(label2, textvariable=DOB, width=150, font=font)
dob_entry.place(x=160, y=100)

CTkLabel(label2, text="Gender:", font=font, ).place(x=30,y=150) #Gender:

#Gender Selection
def radiobutton_event():
    value = radio.get()
    if value == 1:
        gender="Male"
        print (gender)
    else:
        gender="Female"
        print(gender)

radio=IntVar()
R1 = CTkRadioButton(label2, text="Male", variable=radio, value=1, command=radiobutton_event)
R1.place(x=150,y=150)

R2 = CTkRadioButton(label2, text="Female", variable=radio,value=2,command=radiobutton_event)
R2.place(x=250,y=150)
##########

CTkLabel(label2, text="Degree:", font=font, ).place(x=500,y=50)

Degree=StringVar()
Degree_entry = CTkEntry(label2, textvariable=Degree, width=150, font=font)
Degree_entry.place(x=630, y=50)


CTkLabel(label2, text="Major:", font=font, ).place(x=500,y=100)

Major=StringVar()
Major_entry = CTkEntry(label2, textvariable=Major, width=150, font=font)
Major_entry.place(x=630, y=100)


CTkLabel(label2, text="Email Address:", font=font, ).place(x=500,y=150)

Email=StringVar()
Email_entry = CTkEntry(label2, textvariable=Email, width=150, font=font)
Email_entry.place(x=630, y=150)

#Other Details

label3 = CTkLabel(app, font=font,text="", width=900, height=250,  fg_color='#c0c9fe')
label3.place(x=30, y=470)


CTkLabel(label3, text="Contact Number:", font=font, ).place(x=30,y=50)

Contact=StringVar()
Contact_entry = CTkEntry(label3, textvariable=Contact, width=150, font=font)
Contact_entry.place(x=160, y=50)

CTkLabel(label3, text="Address:", font=font, ).place(x=30,y=100)

Address=StringVar()
Address_entry = CTkEntry(label3, textvariable=Address, width=150, font=font)
Address_entry.place(x=160, y=100)

CTkLabel(label3, text="Father's Name:", font=font, ).place(x=500,y=50)

FatherName=StringVar()
FatherName_entry = CTkEntry(label3, textvariable=FatherName, width=150, font=font)
FatherName_entry.place(x=630, y=50)

CTkLabel(label3, text="Previous School:", font=font, ).place(x=500,y=100)

PrvSchool=StringVar()
PrvSchool_entry = CTkEntry(label3, textvariable=PrvSchool, width=150, font=font)
PrvSchool_entry.place(x=630, y=100)

#Image Section

# f = CTkLabel(app, width=200, height=200)
# f.place(x=1000, y=150)

f=CTkFrame(app, width=200, height=200)
f.place(x=1000, y=150)

profile_image = PhotoImage(file="Images/upload holder.png")
profile = CTkLabel(f,text='', image=profile_image)
profile.place(x=0, y=0)

#Buttons------------------->

#Show Image

# def showimage():
#     global filename
#     global img
#     filename = filedialog.askopenfilename(
#         initialdir=os.getcwd(),
#         title="Select the Image File",
#         filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All Files", "*.*"))
#     )

#     img = (Image.open(filename))
#     resized_image= img.resize((190,190))
#     photo2 = ImageTk.PhotoImage(resized_image)
#     f.config(image=photo2)
#     f.image=photo2

# Testing..........


def show_image():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select the Image File",
        filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All Files", "*.*"))
    )

    if filename:        
        img = Image.open(filename)
        resized_image =img.resize((200, 200))
        photo2 = ImageTk.PhotoImage(resized_image)
        label.configure(image=photo2)
        label.image = photo2





label = CTkLabel(f, text='')
label.place(x=0, y=0)

upload_button = CTkButton(app, text="Upload", corner_radius=32, fg_color='#2A409A', hover_color='#c0c9fe', command=show_image)
upload_button.place(x=1000, y=370)


# Testing.......... END



# Ubutton = CTkButton(app, text="Upload", corner_radius=32, fg_color='#2A409A', hover_color='#c0c9fe', command=showimage)
# Ubutton.place(x=1000, y=370)

def Save():
    R1=Registration.get()
    N1=Name.get()
    try:
        G1=radio
    except:
        messagebox.showerror("ERROR!","Please Select Gender")
    # C1=Class.get()

Savebutton = CTkButton(app, text="Save", corner_radius=32, fg_color='green', hover_color='#c0c9fe', command=Save)
Savebutton.place(x=1000, y=450)

#Clear Button 

def clear():
    Name.set('')
    DOB.set('')
    Degree.set('')
    Major.set("")
    Email.set('')
    Contact.set('')
    Address.set('')
    FatherName.set('')
    PrvSchool.set('')
    # radiobutton_event.set(None)

    # Savebutton.configure(state='normal')

    #image resest
    img1= PhotoImage(file='Images/upload holder.png')
    label.configure(image=img1)
    label.image=img1

    
Resetbutton = CTkButton(app, text="Reset", corner_radius=32, fg_color='grey',hover_color='#c0c9fe', command=clear)
Resetbutton.place(x=1000, y=530)

Exitbutton = CTkButton(app, text="Exit", command=exit, corner_radius=32, fg_color='red', hover_color='#c0c9fe' )
Exitbutton.place(x=1000, y=610)

#EXIT Switch
def exit():
    app.destroy()


app.mainloop()
